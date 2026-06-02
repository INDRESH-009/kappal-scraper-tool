"""
Batch auto-search support for Kappal.

This module is intentionally separate from scraper.py so it can be plugged into
server.py later without disturbing the working single-search/export flow.
"""

from __future__ import annotations

import asyncio
import re
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from playwright.async_api import Locator, Page, async_playwright

from scraper import (
    RATES_URL,
    _launch_context,
    _set_section_checkbox,
    _set_service_mode,
    _wait_for_app_idle,
    _wait_for_results_in_context,
    scrape_current_results_page,
)


BATCH_SHEET = "Batch Search Input"

BATCH_COLUMNS = [
    "Origin Port Code",
    "Origin Port Name",
    "Origin Mode",
    "Origin Carrier SD",
    "Origin Include Nearby",
    "Destination Port Code",
    "Destination Port Name",
    "Destination Mode",
    "Destination Carrier SD",
    "Destination Include Nearby",
    "Cut Off Date",
    "Load Type",
    "Quantity",
    "Cargo Weight",
    "Weight Unit",
    "Charges",
    "Stuffing",
    "No Of BL",
    "No Of Shipping Bill",
    "Search Reference Name",
    "Search Currency",
]

DEFAULT_TEST_ROW = {
    "Origin Port Code": "INMAA",
    "Origin Port Name": "Chennai",
    "Origin Mode": "CY",
    "Origin Carrier SD": "No",
    "Origin Include Nearby": "No",
    "Destination Port Code": "USNYC",
    "Destination Port Name": "New York",
    "Destination Mode": "CY",
    "Destination Carrier SD": "No",
    "Destination Include Nearby": "No",
    "Cut Off Date": "02 Jun 2026",
    "Load Type": "20GP",
    "Quantity": 1,
    "Cargo Weight": 17000,
    "Weight Unit": "KG",
    "Charges": "Freight,Origin Charges,Destination Charges",
    "Stuffing": "Factory",
    "No Of BL": 1,
    "No Of Shipping Bill": 1,
    "Search Reference Name": "",
    "Search Currency": "USD",
}


@dataclass
class BatchSearchJob:
    input_row: int
    origin_code: str
    origin_name: str
    origin_mode: str
    origin_carrier_sd: bool
    origin_include_nearby: bool
    destination_code: str
    destination_name: str
    destination_mode: str
    destination_carrier_sd: bool
    destination_include_nearby: bool
    cut_off_date: str
    load_type: str
    quantity: int
    cargo_weight: float
    weight_unit: str
    charges: str
    stuffing: str
    no_of_bl: int
    no_of_shipping_bill: int
    search_reference_name: str
    search_currency: str


@dataclass
class BatchSearchError:
    input_row: int
    origin: str
    destination: str
    load_type: str
    message: str


def create_test_batch_workbook(path: str | Path) -> Path:
    """Creates a one-row input workbook for testing the batch auto-search flow."""
    output_path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = BATCH_SHEET

    header_fill = PatternFill("solid", fgColor="0D1B5E")
    for idx, column in enumerate(BATCH_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(14, min(len(column) + 4, 28))

    for idx, column in enumerate(BATCH_COLUMNS, start=1):
        ws.cell(row=2, column=idx, value=DEFAULT_TEST_ROW.get(column))

    wb.save(output_path)
    return output_path


def parse_batch_workbook(path: str | Path) -> list[BatchSearchJob]:
    wb = load_workbook(path, data_only=True)
    if BATCH_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook must contain a sheet named '{BATCH_SHEET}'.")

    ws = wb[BATCH_SHEET]
    headers = [_clean_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    header_map = {header: idx + 1 for idx, header in enumerate(headers) if header}

    missing = [column for column in BATCH_COLUMNS if column not in header_map]
    if missing:
        raise ValueError(f"Missing required batch columns: {', '.join(missing)}")

    jobs: list[BatchSearchJob] = []
    for row_idx in range(2, ws.max_row + 1):
        raw = {
            column: ws.cell(row_idx, header_map[column]).value
            for column in BATCH_COLUMNS
        }
        if not any(value not in (None, "") for value in raw.values()):
            continue

        jobs.append(
            BatchSearchJob(
                input_row=row_idx,
                origin_code=_required_text(raw, "Origin Port Code").upper(),
                origin_name=_required_text(raw, "Origin Port Name"),
                origin_mode=_mode(raw.get("Origin Mode")),
                origin_carrier_sd=_bool(raw.get("Origin Carrier SD")),
                origin_include_nearby=_bool(raw.get("Origin Include Nearby")),
                destination_code=_required_text(raw, "Destination Port Code").upper(),
                destination_name=_required_text(raw, "Destination Port Name"),
                destination_mode=_mode(raw.get("Destination Mode")),
                destination_carrier_sd=_bool(raw.get("Destination Carrier SD")),
                destination_include_nearby=_bool(raw.get("Destination Include Nearby")),
                cut_off_date=_date_text(raw.get("Cut Off Date")),
                load_type=_required_text(raw, "Load Type").upper(),
                quantity=int(raw.get("Quantity") or 1),
                cargo_weight=float(raw.get("Cargo Weight") or 0),
                weight_unit=str(raw.get("Weight Unit") or "KG").strip().upper(),
                charges=_required_text(raw, "Charges"),
                stuffing=str(raw.get("Stuffing") or "Factory").strip().title(),
                no_of_bl=int(raw.get("No Of BL") or 1),
                no_of_shipping_bill=int(raw.get("No Of Shipping Bill") or 1),
                search_reference_name=str(raw.get("Search Reference Name") or "").strip(),
                search_currency=_required_text(raw, "Search Currency").upper(),
            )
        )

    if not jobs:
        raise ValueError("Batch workbook does not contain any search rows.")
    return jobs


async def run_batch_autosearch(
    workbook_path: str | Path,
    progress_cb: Optional[Callable[[str], Any]] = None,
    headless: bool = False,
) -> dict:
    """
    Runs all searches from a batch workbook and returns a combined scrape payload:
    {"results": [...], "errors": [...], "jobs": [...]}.
    """
    jobs = parse_batch_workbook(workbook_path)
    results: list[dict] = []
    errors: list[BatchSearchError] = []

    async def emit(message: str) -> None:
        if progress_cb:
            await progress_cb(message)

    async with async_playwright() as pw:
        ctx = await _launch_context(pw, headless=headless)
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()

        try:
            await page.goto(RATES_URL, wait_until="domcontentloaded")
            await page.wait_for_load_state("networkidle")
            if "login" in page.url:
                raise RuntimeError("Kappal is not authenticated. Run authentication first.")

            for idx, job in enumerate(jobs, start=1):
                label = f"{job.origin_code}->{job.destination_code} {job.load_type}x{job.quantity}"
                await emit(f"Row {job.input_row} ({idx}/{len(jobs)}): {label} starting")
                try:
                    await _run_one_job(page, job, emit)
                    await emit(f"Row {job.input_row}: search submitted; locating results page")
                    result_target = await _wait_for_results_in_context(ctx, page, 180, emit)
                    await emit(f"Row {job.input_row}: results located; handing off to scraper")
                    row_results = await scrape_current_results_page(result_target, emit)
                    for result in row_results:
                        result["_batch_input_row"] = job.input_row
                        result["_batch_origin_code"] = job.origin_code
                        result["_batch_destination_code"] = job.destination_code
                        result["_batch_load_type"] = job.load_type
                    results.extend(row_results)
                    await emit(f"Row {job.input_row}: complete")
                except Exception as exc:
                    try:
                        await page.screenshot(path=f"debug_batch_row_{job.input_row}_failure.png", full_page=False)
                    except Exception:
                        pass
                    errors.append(
                        BatchSearchError(
                            input_row=job.input_row,
                            origin=job.origin_code,
                            destination=job.destination_code,
                            load_type=job.load_type,
                            message=str(exc),
                        )
                    )
                    await emit(f"Row {job.input_row}: failed - {exc}")
                    await _return_to_search_page(page)

        finally:
            await ctx.close()

    return {
        "search_params": {"mode": "batch_autosearch", "input_file": str(workbook_path)},
        "total_results": len(results),
        "results": results,
        "errors": [asdict(error) for error in errors],
        "jobs": [asdict(job) for job in jobs],
        "scraped_at": datetime.utcnow().isoformat() + "Z",
    }


async def _run_one_job(page: Page, job: BatchSearchJob, emit: Callable[[str], Any]) -> None:
    await _return_to_search_page(page)
    await _reset_form(page)

    await _set_service_mode(page, "origin", job.origin_mode, emit)
    await _set_service_mode(page, "destination", job.destination_mode, emit)
    await _set_section_checkbox(page, "origin", "Carrier SD Services", job.origin_carrier_sd, emit)
    await _set_section_checkbox(page, "origin", "Include Nearby", job.origin_include_nearby, emit)
    await _set_section_checkbox(page, "destination", "Carrier SD Services", job.destination_carrier_sd, emit)
    await _set_section_checkbox(page, "destination", "Include Nearby", job.destination_include_nearby, emit)

    await emit(f"Row {job.input_row}: selecting origin {job.origin_code}")
    await _select_port_exact(page, "origin", job.origin_name, job.origin_code)
    await emit(f"Row {job.input_row}: selecting destination {job.destination_code}")
    await _select_port_exact(page, "destination", job.destination_name, job.destination_code)

    await _select_cutoff_date_from_picker(page, job.cut_off_date)
    await _set_load_type_dialog(page, job.load_type, job.quantity, job.cargo_weight, job.weight_unit)
    await _set_locals_customs_dialog(
        page,
        job.charges,
        job.stuffing,
        job.no_of_bl,
        job.no_of_shipping_bill,
    )
    await _fill_reference_name(page, job.search_reference_name)
    await _select_currency_exact(page, job.search_currency)
    await _click_search_rates(page)


async def _return_to_search_page(page: Page) -> None:
    if "/rates" not in page.url or re.search(r"/rates/(fcl|lcl|air|land)/", page.url, re.I):
        await page.goto(RATES_URL, wait_until="domcontentloaded")
    await _wait_for_app_idle(page)
    await page.wait_for_load_state("networkidle")


async def _reset_form(page: Page) -> None:
    reset = page.locator("text=/^\\s*Reset\\s*$/i").first
    if await reset.count() > 0 and await reset.is_visible():
        await reset.click()
        await asyncio.sleep(0.5)
        await _wait_for_app_idle(page)


async def _select_port_exact(page: Page, section: str, name: str, code: str) -> None:
    field = await _find_port_field(page, section)
    queries = [code, f"{name} {code}", name]
    last_error = ""
    for query in queries:
        try:
            await _clear_and_type(field, query)
            await asyncio.sleep(0.8)
            await _click_dropdown_option_by_code(page, code, field)
            await _wait_for_app_idle(page)
            await asyncio.sleep(0.3)
            if await _port_field_has_code(page, field, code):
                return
            last_error = f"Clicked option for {code}, but selected field did not validate"
        except Exception as exc:
            last_error = str(exc)
        try:
            await _clear_and_type(field, query)
            await asyncio.sleep(0.8)
            await field.press("ArrowDown")
            await asyncio.sleep(0.2)
            await field.press("Enter")
            await _wait_for_app_idle(page)
            await asyncio.sleep(0.3)
            if await _port_field_has_code(page, field, code):
                return
            last_error = f"Keyboard-selected option for {code}, but field did not validate"
        except Exception as exc:
            last_error = str(exc)
    try:
        await page.screenshot(path=f"debug_batch_{section}_{code}.png", full_page=False)
    except Exception:
        pass
    raise RuntimeError(f"Could not select {section} port {name} / {code}: {last_error}")


async def _find_port_field(page: Page, section: str) -> Locator:
    selectors = (
        [
            "input[name='fcl_origin_port']:visible",
            "input#originLocation:visible",
            "md-input-container[md-input-id='originLocation'] input:visible",
        ]
        if section == "origin"
        else [
            "input[name='fcl_destination_port']:visible",
            "input#destinationLocation:visible",
            "md-input-container[md-input-id='destinationLocation'] input:visible",
        ]
    )
    for selector in selectors:
        locator = page.locator(selector).first
        if await locator.count() > 0 and await locator.is_visible():
            return locator

    label = "Origin" if section == "origin" else "Destination"
    field = page.locator(
        f"xpath=//*[normalize-space()='{label}']/following::*[self::input or self::div][1]"
    ).first
    if await field.count() == 0:
        raise RuntimeError(f"Could not find {section} port field")
    return field


async def _clear_and_type(locator: Locator, value: str) -> None:
    await locator.click(force=True)
    await locator.fill("")
    await locator.type(value, delay=60)


async def _click_dropdown_option_by_code(page: Page, code: str, field: Locator) -> None:
    field_box = await field.bounding_box()
    if not field_box:
        raise RuntimeError(f"Could not read port field position for {code}")

    clicked = await page.evaluate(
        """
        ({code, fieldBox}) => {
            const visible = (el) => {
                const rect = el.getBoundingClientRect();
                const style = getComputedStyle(el);
                return rect.width > 0 && rect.height > 0
                    && style.display !== 'none'
                    && style.visibility !== 'hidden';
            };

            const belowField = (el) => {
                const rect = el.getBoundingClientRect();
                return rect.top >= fieldBox.y + fieldBox.height - 8
                    && rect.top <= fieldBox.y + fieldBox.height + 420
                    && rect.left >= fieldBox.x - 40
                    && rect.left <= fieldBox.x + fieldBox.width + 80;
            };

            const makeRow = (el) => {
                let row = el;
                for (let i = 0; i < 6 && row && row.parentElement; i++) {
                    const rect = row.getBoundingClientRect();
                    const text = row.innerText || row.textContent || '';
                    if (
                        text.includes(code)
                        && belowField(row)
                        && rect.width >= Math.min(fieldBox.width * 0.35, 260)
                        && rect.height >= 28
                        && rect.height <= 120
                    ) {
                        return row;
                    }
                    row = row.parentElement;
                }
                return el;
            };

            const all = Array.from(document.querySelectorAll(
                '[role="option"], md-autocomplete-parent-scope, md-virtual-repeat-container *, [role="listbox"] *, ul li, div, span, button'
            ));
            const candidates = all
                .filter(el => {
                    const text = el.innerText || el.textContent || '';
                    if (!visible(el) || !text.includes(code) || !belowField(el)) return false;
                    const rect = el.getBoundingClientRect();
                    if (rect.height < 12 || rect.height > 160) return false;
                    if (rect.width < 35 || rect.width > fieldBox.width + 120) return false;
                    if (el.tagName === 'INPUT') return false;
                    return true;
                })
                .map(makeRow)
                .filter((el, idx, arr) => arr.indexOf(el) === idx)
                .sort((a, b) => {
                    const ar = a.getBoundingClientRect();
                    const br = b.getBoundingClientRect();
                    return ar.top - br.top || br.width - ar.width;
                });

            const target = candidates[0];
            if (!target) return false;
            const rect = target.getBoundingClientRect();
            target.dispatchEvent(new MouseEvent('mousedown', {bubbles: true, clientX: rect.left + rect.width / 2, clientY: rect.top + rect.height / 2}));
            target.dispatchEvent(new MouseEvent('mouseup', {bubbles: true, clientX: rect.left + rect.width / 2, clientY: rect.top + rect.height / 2}));
            target.click();
            return true;
        }
        """,
        {"code": code, "fieldBox": field_box},
    )
    if not clicked:
        raise RuntimeError(f"Dropdown option with code {code} not found")


async def _port_field_has_code(page: Page, field: Locator, code: str) -> bool:
    field_box = await field.bounding_box()
    if not field_box:
        return False
    return bool(await page.evaluate(
        """
        ({code, fieldBox}) => {
            const visible = (el) => {
                const rect = el.getBoundingClientRect();
                const style = getComputedStyle(el);
                return rect.width > 0 && rect.height > 0
                    && style.display !== 'none'
                    && style.visibility !== 'hidden';
            };
            const nearField = (el) => {
                const rect = el.getBoundingClientRect();
                return rect.top >= fieldBox.y - 30
                    && rect.top <= fieldBox.y + fieldBox.height + 40
                    && rect.left >= fieldBox.x - 30
                    && rect.left <= fieldBox.x + fieldBox.width + 30;
            };
            const candidates = Array.from(document.querySelectorAll('input, div, span, md-input-container'))
                .filter(el => visible(el) && nearField(el));
            const hasCode = candidates.some(el => ((el.innerText || el.textContent || el.value || '').toUpperCase()).includes(code));
            if (!hasCode) return false;
            const redInvalid = candidates.some(el => {
                const cls = `${el.className || ''}`.toLowerCase();
                if (cls.includes('ng-invalid')) return true;
                const border = getComputedStyle(el).borderColor;
                return /rgb\\(255,\\s*87,\\s*34\\)|rgb\\(244,\\s*67,\\s*54\\)|red/i.test(border);
            });
            return !redInvalid;
        }
        """,
        {"code": code.upper(), "fieldBox": field_box},
    ))


async def _port_field_root(page: Page, section: str) -> Locator:
    field = await _find_port_field(page, section)
    root = field.locator("xpath=ancestor::*[contains(@class,'input') or contains(@class,'port') or contains(@class,'location')][1]")
    if await root.count() > 0:
        return root.first
    return field


async def _select_cutoff_date_from_picker(page: Page, date_text: str) -> None:
    target = _parse_date(date_text)
    target_display = target.strftime("%d %b %Y")
    if await _visible_text_exists(page, target_display):
        return

    date_trigger = page.locator(
        ".md-datepicker-input:visible, input.md-datepicker-input:visible, "
        "button:has-text('calendar'), [class*='datepicker']:visible, "
        "div:has-text('Cut Off Date') >> xpath=following::*[contains(@class,'date') or self::input][1]"
    ).first
    if await date_trigger.count() == 0:
        raise RuntimeError("Cut Off Date picker field not found")
    await date_trigger.click(force=True)
    await asyncio.sleep(0.3)

    for _ in range(24):
        visible_text = await page.locator("body").inner_text()
        wanted_month = target.strftime("%b %Y")
        long_month = target.strftime("%B %Y")
        if wanted_month in visible_text or long_month in visible_text:
            break
        next_button = page.locator(
            "button[aria-label*='Next' i], .md-datepicker-next, "
            "button:has-text('›'), button:has-text('>')"
        ).first
        if await next_button.count() == 0:
            raise RuntimeError(f"Could not navigate date picker to {wanted_month}")
        await next_button.click(force=True)
        await asyncio.sleep(0.2)

    clicked = await page.evaluate(
        """
        (day) => {
            const visible = (el) => {
                const rect = el.getBoundingClientRect();
                const style = getComputedStyle(el);
                return rect.width > 0 && rect.height > 0
                    && style.display !== 'none'
                    && style.visibility !== 'hidden';
            };
            const candidates = Array.from(document.querySelectorAll(
                'button, td, [role="gridcell"], .md-calendar-date, .md-calendar-date-selection-indicator'
            )).filter(el => visible(el) && (el.innerText || el.textContent || '').trim() === String(day));
            const target = candidates
                .filter(el => {
                    const rect = el.getBoundingClientRect();
                    return rect.width <= 80 && rect.height <= 80;
                })
                .sort((a, b) => a.getBoundingClientRect().top - b.getBoundingClientRect().top)[0];
            if (!target) return false;
            target.click();
            return true;
        }
        """,
        target.day,
    )
    if not clicked:
        try:
            await page.screenshot(path="debug_batch_date.png", full_page=False)
        except Exception:
            pass
        raise RuntimeError(f"Date picker day {target.day} not found")
    await asyncio.sleep(0.3)


async def _visible_text_exists(page: Page, text: str) -> bool:
    return bool(await page.evaluate(
        """
        (text) => {
            const visible = (el) => {
                const rect = el.getBoundingClientRect();
                const style = getComputedStyle(el);
                return rect.width > 0 && rect.height > 0
                    && style.display !== 'none'
                    && style.visibility !== 'hidden';
            };
            return Array.from(document.querySelectorAll('input, div, span, button'))
                .some(el => visible(el) && ((el.innerText || el.textContent || el.value || '').trim()).includes(text));
        }
        """,
        text,
    ))


async def _set_load_type_dialog(
    page: Page,
    load_type: str,
    quantity: int,
    cargo_weight: float,
    weight_unit: str,
) -> None:
    trigger = page.locator("text=/Load Type/i").locator("xpath=following::*[contains(text(),'GP') or contains(text(),'HC') or self::button or self::div][1]").first
    if await trigger.count() == 0:
        trigger = page.locator("div:has-text('20GP'), div:has-text('Load Type')").first
    await trigger.click(force=True)
    await asyncio.sleep(0.4)

    if not await _visible_text_exists(page, load_type):
        await _select_dialog_dropdown_value(page, load_type)
    await _fill_load_quantity_and_weight(page, quantity, cargo_weight)
    if weight_unit.upper() != "KG":
        await _select_dialog_dropdown_value(page, weight_unit)

    done = page.locator("button:has-text('Done'), div:has-text('Done')").last
    if await done.count() == 0:
        raise RuntimeError("Load Type dialog Done button not found")
    await done.click(force=True)
    await asyncio.sleep(0.4)
    if not await _visible_text_exists(page, f"{load_type} x{quantity}"):
        if not await _visible_text_exists(page, f"{load_type}x{quantity}"):
            try:
                await page.screenshot(path="debug_batch_load_type.png", full_page=False)
            except Exception:
                pass
            raise RuntimeError(f"Load type summary did not validate as {load_type} x{quantity}")


async def _set_locals_customs_dialog(
    page: Page,
    charges: str,
    stuffing: str,
    no_of_bl: int,
    no_of_shipping_bill: int,
) -> None:
    trigger = page.locator("text=/Locals\\s*&\\s*Custom Charges/i").locator("xpath=following::*[self::div or self::button][1]").first
    if await trigger.count() == 0:
        trigger = page.locator("div:has-text('Freight'), div:has-text('Origin')").first
    await trigger.click(force=True)
    await asyncio.sleep(0.4)

    for option in _charge_options(charges):
        await _set_named_checkbox(page, option, True)

    await _select_radio(page, stuffing)
    await _fill_near_label(page, "No. Of B/L", str(no_of_bl))
    await _fill_near_label(page, "No. Of Shipping Bill", str(no_of_shipping_bill))

    done = page.locator("button:has-text('Done'), div:has-text('Done')").last
    if await done.count() == 0:
        raise RuntimeError("Locals and Customs dialog Done button not found")
    await done.click(force=True)
    await asyncio.sleep(0.4)


async def _fill_load_quantity_and_weight(page: Page, quantity: int, cargo_weight: float) -> None:
    ok = await page.evaluate(
        """
        ({quantity, cargoWeight}) => {
            const visible = (el) => {
                const rect = el.getBoundingClientRect();
                const style = getComputedStyle(el);
                return rect.width > 0 && rect.height > 0
                    && style.display !== 'none'
                    && style.visibility !== 'hidden';
            };
            const textOf = (el) => el ? (el.innerText || el.textContent || '').trim() : '';
            const dialogs = Array.from(document.querySelectorAll('div, md-card, md-dialog, section'))
                .filter(el => {
                    if (!visible(el)) return false;
                    const text = textOf(el);
                    return text.includes('Load Type')
                        && text.includes('Quantity')
                        && text.includes('Cargo Weight')
                        && text.includes('Done');
                })
                .sort((a, b) => {
                    const ar = a.getBoundingClientRect();
                    const br = b.getBoundingClientRect();
                    return (ar.width * ar.height) - (br.width * br.height);
                });
            const dialog = dialogs[0];
            if (!dialog) return {ok: false, reason: 'load dialog not found'};

            const inputs = Array.from(dialog.querySelectorAll('input'))
                .filter(visible)
                .sort((a, b) => {
                    const ar = a.getBoundingClientRect();
                    const br = b.getBoundingClientRect();
                    return ar.top - br.top || ar.left - br.left;
                });
            if (inputs.length < 2) return {ok: false, reason: `expected 2 visible inputs, found ${inputs.length}`};

            const setValue = (input, value) => {
                const proto = Object.getPrototypeOf(input);
                const desc = Object.getOwnPropertyDescriptor(proto, 'value');
                if (desc && desc.set) desc.set.call(input, String(value));
                else input.value = String(value);
                input.dispatchEvent(new Event('input', {bubbles: true}));
                input.dispatchEvent(new Event('change', {bubbles: true}));
                input.dispatchEvent(new Event('blur', {bubbles: true}));
            };

            setValue(inputs[0], quantity);
            setValue(inputs[1], cargoWeight);
            return {ok: true, values: inputs.map(i => i.value)};
        }
        """,
        {"quantity": str(quantity), "cargoWeight": _number_text(cargo_weight)},
    )
    if not ok or not ok.get("ok"):
        try:
            await page.screenshot(path="debug_batch_load_dialog.png", full_page=False)
        except Exception:
            pass
        reason = ok.get("reason") if isinstance(ok, dict) else "unknown"
        raise RuntimeError(f"Could not fill load quantity/weight: {reason}")


async def _select_dialog_dropdown_value(page: Page, value: str) -> None:
    option = page.locator(
        f"[role='option']:has-text('{value}'), li:has-text('{value}'), "
        f"md-option:has-text('{value}'), div[class*='option']:has-text('{value}')"
    )
    count = await option.count()
    for idx in range(min(count, 12)):
        candidate = option.nth(idx)
        if await candidate.is_visible():
            await candidate.click(force=True)
            return

    trigger = page.locator(f"div:has-text('{value}'), button:has-text('{value}'), md-select:has-text('{value}')").first
    if await trigger.count() > 0:
        await trigger.click(force=True)
        await asyncio.sleep(0.2)
        option = page.locator(
            f"[role='option']:has-text('{value}'), li:has-text('{value}'), md-option:has-text('{value}')"
        )
        count = await option.count()
        for idx in range(min(count, 12)):
            candidate = option.nth(idx)
            if await candidate.is_visible():
                await candidate.click(force=True)
                return
    raise RuntimeError(f"Could not select dropdown value {value}")


async def _fill_near_label(page: Page, label: str, value: str) -> None:
    locator = page.locator(
        f"xpath=//*[contains(normalize-space(),'{label}')]/following::input[1]"
    ).first
    if await locator.count() == 0:
        raise RuntimeError(f"Input for '{label}' not found")
    await locator.fill("")
    await locator.type(value, delay=20)


async def _set_named_checkbox(page: Page, label: str, should_check: bool) -> None:
    box = page.locator(
        f"md-checkbox:has-text('{label}'), [role='checkbox']:has-text('{label}'), label:has-text('{label}')"
    ).first
    if await box.count() == 0:
        return
    checked = (await box.get_attribute("aria-checked")) == "true"
    if checked != should_check:
        await box.click(force=True)
        await asyncio.sleep(0.1)


async def _select_radio(page: Page, label: str) -> None:
    radio = page.locator(
        f"md-radio-button:has-text('{label}'), [role='radio']:has-text('{label}'), label:has-text('{label}')"
    ).first
    if await radio.count() > 0:
        await radio.click(force=True)


async def _fill_reference_name(page: Page, reference_name: str) -> None:
    if not reference_name:
        return
    ref = page.locator(
        "input[placeholder*='reference' i]:visible, input[name*='reference' i]:visible"
    ).first
    if await ref.count() > 0:
        await ref.fill(reference_name)


async def _select_currency_exact(page: Page, currency: str) -> None:
    trigger = page.locator(
        "text=/Search Currency/i"
    ).locator("xpath=following::*[self::div or self::button or self::md-select][1]").first
    if await trigger.count() == 0:
        trigger = page.locator("md-select:has-text('USD'), div:has-text('USD'), button:has-text('USD')").last
    await trigger.click(force=True)
    await asyncio.sleep(0.3)

    option = page.locator(
        f"[role='option']:has-text('{currency}'), md-option:has-text('{currency}'), "
        f"li:has-text('{currency}'), div[class*='option']:has-text('{currency}')"
    ).first
    if await option.count() == 0:
        raise RuntimeError(f"Currency option {currency} not found")
    await option.click(force=True)
    await asyncio.sleep(0.2)


async def _click_search_rates(page: Page) -> None:
    button = page.locator("button:has-text('Search Rates'), a:has-text('Search Rates')").last
    if await button.count() == 0:
        raise RuntimeError("Search Rates button not found")
    await button.click(force=True)
    await page.wait_for_load_state("domcontentloaded")
    await _wait_for_app_idle(page)


def _charge_options(charges: str) -> list[str]:
    normalized = charges.lower()
    options = []
    if "freight" in normalized:
        options.append("Freight")
    if "origin" in normalized:
        options.append("Origin Charges")
    if "destination" in normalized:
        options.append("Destination Charges")
    if "origin custom" in normalized:
        options.append("Origin Custom Charges")
    if "destination custom" in normalized:
        options.append("Destination Custom Charges")
    return options


def _clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _required_text(row: dict[str, Any], key: str) -> str:
    value = row.get(key)
    if value in (None, ""):
        raise ValueError(f"Missing required value for '{key}'.")
    return str(value).strip()


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"yes", "y", "true", "1", "on"}


def _mode(value: Any) -> str:
    mode = str(value or "CY").strip().upper()
    if mode not in {"CY", "DOOR"}:
        raise ValueError(f"Mode must be CY or DOOR, got {mode!r}.")
    return mode


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")
    if isinstance(value, date):
        return value.strftime("%d %b %Y")
    return str(value or "").strip()


def _parse_date(value: str) -> date:
    for fmt in ("%d %b %Y", "%d %B %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date format: {value!r}")


def _number_text(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else str(value)


if __name__ == "__main__":
    created = create_test_batch_workbook(Path("batch_search_test_input.xlsx"))
    print(f"Created {created}")
