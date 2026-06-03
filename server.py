"""
server.py
─────────
FastAPI backend for kappal auto scrapper.
• Serves the frontend at  GET  /
• Accepts scrape jobs via  WebSocket  /ws/scrape
  – Streams progress messages back in real time
  – Sends the final JSON result when done
"""

import json
import asyncio
import re
import uuid
import base64
from io import BytesIO
from pathlib import Path

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from scraper import authenticate_kappal, manual_search_and_scrape_kappal, scrape_kappal
from batch_autosearch import (
    BATCH_SHEET,
    create_test_batch_workbook,
    parse_batch_workbook,
    run_batch_autosearch,
)

# ─── App setup ────────────────────────────────────────────────────────────────

app = FastAPI(title="kappal auto scrapper", version="1.0.0")

STATIC_DIR = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
BATCH_UPLOAD_DIR = Path(__file__).parent / "batch_uploads"
BATCH_UPLOAD_DIR.mkdir(exist_ok=True)
BATCH_UPLOADS: dict[str, Path] = {}


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def root():
    return (STATIC_DIR / "index.html").read_text()


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.post("/export/excel")
async def export_excel(payload: dict):
    workbook = build_rates_workbook(payload)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = "kappal_rates.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/batch/template")
async def batch_template():
    output = BytesIO()
    workbook_path = BATCH_UPLOAD_DIR / "batch_search_test_input.xlsx"
    create_test_batch_workbook(workbook_path)
    with workbook_path.open("rb") as f:
        output.write(f.read())
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="batch_search_test_input.xlsx"'},
    )


@app.post("/batch/upload")
async def batch_upload(payload: dict):
    filename = str(payload.get("filename") or "")
    content_base64 = str(payload.get("content_base64") or "")
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        return JSONResponse({"error": "Upload must be an .xlsx or .xlsm workbook."}, status_code=400)
    if not content_base64:
        return JSONResponse({"error": "Missing workbook content."}, status_code=400)

    upload_id = uuid.uuid4().hex
    safe_name = Path(filename).name
    path = BATCH_UPLOAD_DIR / f"{upload_id}_{safe_name}"
    try:
        path.write_bytes(base64.b64decode(content_base64))
    except Exception:
        return JSONResponse({"error": "Workbook content is not valid base64."}, status_code=400)

    try:
        jobs = parse_batch_workbook(path)
    except Exception as exc:
        path.unlink(missing_ok=True)
        return JSONResponse({"error": str(exc)}, status_code=400)

    BATCH_UPLOADS[upload_id] = path
    return {
        "upload_id": upload_id,
        "filename": safe_name,
        "sheet": BATCH_SHEET,
        "jobs": len(jobs),
        "first_job": jobs[0].__dict__ if jobs else None,
    }


# ─── WebSocket scrape endpoint ────────────────────────────────────────────────

class ScrapeRequest(BaseModel):
    origin:           str
    destination:      str
    cut_off_date:     str
    load_type:        str  = "20GP"
    quantity:         int  = 1
    origin_service_mode: str = "CY"
    destination_service_mode: str = "CY"
    origin_carrier_sd: bool = False
    destination_carrier_sd: bool = False
    include_nearby_origin: bool = False
    include_nearby_destination: bool = False
    charges: str = "Freight, Origin, Destination +1"
    search_reference_name: str = ""
    search_currency: str = "USD"


class AuthRequest(BaseModel):
    auth_timeout: int = 300   # seconds to wait for manual login/CAPTCHA solve


class ManualScrapeRequest(BaseModel):
    search_timeout: int = 600   # seconds to wait for the user to submit search


# ─── Template column structure (matches Rate_Capture_Template.xlsx exactly) ──

# Header color scheme (from template)
COLOR_RED       = "DC2626"   # Required columns
COLOR_NAVY      = "0D1B5E"   # Route metadata
COLOR_BLUE      = "1E3A8A"   # FC (Freight Charges)
COLOR_BROWN     = "92400E"   # OC (Origin Charges)
COLOR_GREEN     = "065F46"   # DC (Destination Charges)
COLOR_ROW_BG    = "F0F4FF"   # Alternating data row background

# Required columns (red header)
REQUIRED_COLS = {"Mode", "Shipping Line", "POL Code", "POD Code", "Valid From"}

# Route metadata columns (first 22)
ROUTE_COLS = [
    "Mode", "Rate Type", "Shipping Line", "Shipping Line Code", "Container Type",
    "Service Mode", "Service Name", "POL Code", "POL Name", "Origin Terminal",
    "POD Code", "POD Name", "Destination Terminal", "Via Codes", "Via Names",
    "Sailing Date", "Transit Days", "Free Days", "Cargo Type", "Cargo Description",
    "Valid From", "Valid To",
]

# Build full 144-column header list
def _build_header() -> list[str]:
    headers = list(ROUTE_COLS)
    for i in range(1, 7):   # FC1–FC6
        for f in ("Name", "Code", "Basis", "Currency", "Amount"):
            headers.append(f"FC{i} {f}")
    for i in range(1, 10):  # OC1–OC9
        for f in ("Name", "Code", "Basis", "Currency", "Amount"):
            headers.append(f"OC{i} {f}")
    for i in range(1, 10):  # DC1–DC9
        for f in ("Name", "Code", "Basis", "Currency", "Amount"):
            headers.append(f"DC{i} {f}")
    headers.append("Inclusions")
    headers.append("Remarks")
    return headers

HEADERS = _build_header()  # 144 columns


def _header_color(col_name: str) -> str:
    if col_name in REQUIRED_COLS:
        return COLOR_RED
    if col_name.startswith("FC"):
        return COLOR_BLUE
    if col_name.startswith("OC"):
        return COLOR_BROWN
    if col_name.startswith("DC"):
        return COLOR_GREEN
    return COLOR_NAVY


def build_rates_workbook(payload: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"

    # ── Write header row ──────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=_header_color(col_name))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Write data rows ───────────────────────────────────────────────────────
    exportable_rates = [rate for rate in payload.get("results") or [] if _is_exportable_rate(rate)]
    for row_idx, rate in enumerate(exportable_rates, start=2):
        row_data = _rate_to_template_row(rate)
        use_bg = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            if use_bg:
                cell.fill = PatternFill("solid", fgColor=COLOR_ROW_BG)

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "Mode": 10, "Rate Type": 12, "Shipping Line": 18, "Shipping Line Code": 10,
        "Container Type": 12, "Service Mode": 10, "Service Name": 14,
        "POL Code": 9, "POL Name": 22, "Origin Terminal": 16,
        "POD Code": 9, "POD Name": 22, "Destination Terminal": 16,
        "Via Codes": 12, "Via Names": 18,
        "Sailing Date": 13, "Transit Days": 11, "Free Days": 9,
        "Cargo Type": 10, "Cargo Description": 16,
        "Valid From": 12, "Valid To": 12,
        "Inclusions": 30, "Remarks": 30,
    }
    for col_idx, col_name in enumerate(HEADERS, start=1):
        letter = get_column_letter(col_idx)
        width = col_widths.get(col_name, 13 if "Name" in col_name else 10)
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    _add_errors_sheet(wb, payload.get("errors") or [])

    return wb


def _add_errors_sheet(wb: Workbook, errors: list[dict]) -> None:
    if not errors:
        return
    ws = wb.create_sheet("Errors")
    headers = ["Input Row", "Origin", "Destination", "Load Type", "Message"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=COLOR_RED)
    for row_idx, error in enumerate(errors, start=2):
        ws.cell(row=row_idx, column=1, value=error.get("input_row"))
        ws.cell(row=row_idx, column=2, value=error.get("origin"))
        ws.cell(row=row_idx, column=3, value=error.get("destination"))
        ws.cell(row=row_idx, column=4, value=error.get("load_type"))
        ws.cell(row=row_idx, column=5, value=error.get("message"))
    widths = [12, 12, 14, 12, 80]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def _is_exportable_rate(rate: dict) -> bool:
    if not isinstance(rate, dict) or rate.get("_error"):
        return False
    has_route = bool(
        rate.get("carrier")
        or rate.get("card_carrier")
        or rate.get("port_of_loading")
        or rate.get("port_of_origin")
        or rate.get("port_of_discharge")
    )
    charges = rate.get("charges") or {}
    has_charges = any(
        _get_items(charges, section)
        for section in ("freight", "origin", "destination")
    )
    return has_route or has_charges


def _rate_to_template_row(rate: dict) -> list:
    """Converts a scraped rate dict to the 144-column template row."""
    charges     = rate.get("charges") or {}
    freight_items   = _get_items(charges, "freight")
    origin_items    = _get_items(charges, "origin")
    dest_items      = _get_items(charges, "destination")
    remarks_block   = rate.get("remarks_and_inclusions") or {}

    # ── Sailing / validity dates ──────────────────────────────────────────────
    sailing = rate.get("sailing_date") or rate.get("card_sailing_date") or ""

    # Use the Effective Period parsed directly from the modal header first
    valid_from = rate.get("valid_from") or _extract_valid_from(rate)
    valid_to   = rate.get("valid_to")   or _extract_valid_to(rate)

    # Normalise dates to YYYY-MM-DD
    valid_from = _normalise_date(valid_from)
    valid_to   = _normalise_date(valid_to)
    sailing    = _normalise_date(sailing) or sailing  # keep original if parse fails

    # ── Transit days (strip " Days" suffix) ──────────────────────────────────
    transit_raw = rate.get("transit_time") or rate.get("card_transit_time") or ""
    transit_days = re.search(r"\d+", str(transit_raw))
    transit_days = int(transit_days.group()) if transit_days else None

    # ── Route metadata (22 cols) ──────────────────────────────────────────────
    carrier = rate.get("carrier") or rate.get("card_carrier") or ""
    row = [
        "SEA-FCL",
        "SPOT RATE",
        carrier,
        _carrier_code(carrier),
        _container_type(rate),
        _service_mode_str(rate),
        rate.get("service_name"),
        rate.get("port_of_loading") or rate.get("port_of_origin"),
        _pol_name(rate),
        None,                                              # Origin Terminal
        rate.get("port_of_discharge"),
        _pod_name(rate),
        None,                                              # Destination Terminal
        rate.get("transshipment_port"),
        None,                                              # Via Names
        sailing or None,
        transit_days,
        None,                                              # Free Days
        rate.get("cargo_type") or rate.get("card_cargo_type") or "FAK",
        None,                                              # Cargo Description
        valid_from,
        valid_to,
    ]

    # ── FC1–FC6 (30 cols) ─────────────────────────────────────────────────────
    for i in range(6):
        row.extend(_charge_cols(freight_items, i))

    # ── OC1–OC9 (45 cols) ────────────────────────────────────────────────────
    for i in range(9):
        row.extend(_charge_cols(origin_items, i))

    # ── DC1–DC9 (45 cols) ────────────────────────────────────────────────────
    for i in range(9):
        row.extend(_charge_cols(dest_items, i))

    # ── Inclusions, Remarks (2 cols) ─────────────────────────────────────────
    row.append(remarks_block.get("inclusions") or None)
    row.append(remarks_block.get("remarks") or None)

    return row


def _get_items(charges: dict, section: str) -> list:
    sec = charges.get(section)
    if not isinstance(sec, dict):
        return []
    items = []
    seen = set()
    for item in sec.get("line_items") or []:
        name = _clean_charge_name((item or {}).get("name"))
        if not _looks_like_charge_name(name):
            continue
        ident = name.lower()
        if ident in seen:
            continue
        seen.add(ident)
        item = dict(item)
        item["name"] = name
        items.append(item)
    return items


def _charge_cols(items: list, idx: int) -> list:
    """Returns [Name, Code, Basis, Currency, Amount] for item at idx, or 5 Nones."""
    if idx >= len(items):
        return [None, None, None, None, None]
    item = items[idx]
    amt = _amount_parts(item.get("amount"))
    return [
        item.get("name"),
        _charge_code(item.get("name")),
        item.get("basis"),
        amt[0],
        amt[1],
    ]


# ─── Helpers ──────────────────────────────────────────────────────────────────

KNOWN_CARRIER_CODES = {
    "oocl": "OOCL", "maersk": "MAEU", "msc": "MSCU",
    "hapag": "HLCU", "one": "ONE", "cosco": "COSU",
    "yang ming": "YMLU", "evergreen": "EGLV", "cma": "CMDU",
    "zim": "ZIMU", "pil": "PABV", "wan hai": "WHLC",
}

KNOWN_CHARGE_CODES = {
    "basic ocean freight": "BOF", "marine fuel recovery": "MFR",
    "emergency fuel surcharge": "EFS", "carrier security surcharge": "CSS",
    "security manifest document fee": "SMDF", "document charge": "DOC",
    "export service fee": "ESF", "origin terminal handling charge": "OTHC",
    "terminal handling charge": "THC", "destination terminal handling charge": "DTHC",
    "ny pass through charge": "NYPT", "new york pass through": "NYPT",
    "cfs fee": "CFS", "heavy weight fee": "HWF", "special dimension fee": "SDF",
    "fuel bunker fee": "FBF", "terminal security charges": "TSC",
    "equipment maintenance fee": "EMF", "isps": "ISPS",
    "low sulphur surcharge": "LSS", "peak season surcharge": "PSS",
    "congestion surcharge": "CGS", "war risk surcharge": "WRS",
}

KNOWN_PORT_NAMES = {
    "AUMEL": "Melbourne",
    "USNYC": "New York, NY",
}


def _carrier_code(name: str | None) -> str | None:
    if not name:
        return None
    low = name.lower()
    for key, code in KNOWN_CARRIER_CODES.items():
        if key in low:
            return code
    return (name.split()[0].upper())[:6]


def _charge_code(name: str | None) -> str | None:
    if not name:
        return None
    cleaned = re.sub(r"\s+\d{2}(GP|HC|HQ|DV|RF|OT|RE)\s*$", "", name.strip(), flags=re.I)
    paren_code = re.search(r"\(([A-Z]{2,6})\)", cleaned)
    if paren_code:
        return paren_code.group(1)
    cleaned = re.sub(r"\s*\([^)]*\)", "", cleaned).strip().lower()
    return KNOWN_CHARGE_CODES.get(cleaned)


def _clean_charge_name(value: str | None) -> str:
    name = str(value or "").replace("\t", " ")
    return re.sub(r"\s+", " ", name).strip()


def _looks_like_charge_name(value: str | None) -> bool:
    name = _clean_charge_name(value)
    low = name.lower()
    if not name or low in {"undefined", "free days", "total"}:
        return False
    if len(name) < 4 or len(name) > 90:
        return False
    if not re.search(r"[A-Za-z]", name):
        return False
    if re.fullmatch(r"[A-Z]{3}", name):
        return False
    if re.fullmatch(r"[\d,]+(?:\.\d+)?", name):
        return False
    if re.fullmatch(r"\d{2}\s*(GP|HC|HQ|DV|RF|OT|RE)", name, re.I):
        return False
    if re.fullmatch(r"per(?:\s+\w+){0,3}", name, re.I):
        return False
    if low in {"charges", "basis", "equipment type", "quantity", "quantity | slab", "unit price", "amount", "comments", "sub total", "total cost"}:
        return False
    if "basis equipment type" in low or "unit price amount" in low:
        return False
    return True


def _container_type(rate: dict) -> str | None:
    for section in (rate.get("charges") or {}).values():
        if not isinstance(section, dict):
            continue
        for item in section.get("line_items") or []:
            eq = item.get("equipment_type")
            if eq:
                return eq.upper()
    return None


def _service_mode_str(rate: dict) -> str | None:
    o = rate.get("origin_service_mode")
    d = rate.get("destination_service_mode")
    if o and d:
        return f"{o}/{d}"
    return rate.get("service_type") or rate.get("card_service_type")


def _pol_name(rate: dict) -> str | None:
    """Best-effort port of loading name from scraped data."""
    name = rate.get("pol_name") or rate.get("port_of_loading_name")
    if _looks_like_port_name(name):
        return name
    return KNOWN_PORT_NAMES.get(rate.get("port_of_loading") or rate.get("port_of_origin"))


def _pod_name(rate: dict) -> str | None:
    name = rate.get("pod_name") or rate.get("port_of_discharge_name")
    if _looks_like_port_name(name):
        return name
    return KNOWN_PORT_NAMES.get(rate.get("port_of_discharge"))


def _looks_like_port_name(value: str | None) -> bool:
    name = (value or "").strip()
    if not name:
        return False
    if re.fullmatch(r"S\s+[A-Z]{5}", name):
        return False
    if re.fullmatch(r"[A-Z]{5}", name):
        return False
    return True


def _normalise_date(date_str: str | None) -> str | None:
    """Converts '13 Jun 2026' or '2026-06-13' → '2026-06-13'. Returns None if unparseable."""
    if not date_str:
        return None
    from datetime import datetime
    for fmt in ("%d %b %Y", "%d %B %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return date_str  # return as-is if we can't parse


def _extract_valid_from(rate: dict) -> str | None:
    """Tries to pull a Valid From date from remarks or card fields."""
    remarks = (rate.get("remarks_and_inclusions") or {}).get("remarks") or ""
    m = re.search(r"(\d{2}\s+[A-Za-z]{3}\s+\d{4})", remarks)
    if m:
        try:
            from datetime import datetime
            return datetime.strptime(m.group(1), "%d %b %Y").strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def _extract_valid_to(rate: dict) -> str | None:
    remarks = (rate.get("remarks_and_inclusions") or {}).get("remarks") or ""
    dates = re.findall(r"(\d{2}\s+[A-Za-z]{3}\s+\d{4})", remarks)
    if len(dates) >= 2:
        try:
            from datetime import datetime
            return datetime.strptime(dates[1], "%d %b %Y").strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def _amount_parts(value) -> tuple:
    if isinstance(value, dict):
        return value.get("currency"), value.get("amount")
    if isinstance(value, (int, float)):
        return "USD", value
    if isinstance(value, str):
        m = re.search(r"([A-Z]{3})?\s*([\d,]+(?:\.\d+)?)", value)
        if m:
            return m.group(1), float(m.group(2).replace(",", ""))
    return None, None


@app.websocket("/ws/authenticate")
async def ws_authenticate(ws: WebSocket):
    await ws.accept()

    async def send(obj: dict):
        await ws.send_text(json.dumps(obj))

    try:
        raw = await ws.receive_text()
        req = AuthRequest(**json.loads(raw))

        async def progress(msg: str):
            await send({"type": "progress", "message": msg})

        result = await authenticate_kappal(
            progress_cb=progress,
            headless=False,
            auth_timeout=req.auth_timeout,
        )

        await send({"type": "complete", "data": result})

    except WebSocketDisconnect:
        pass
    except json.JSONDecodeError as e:
        await send({"type": "error", "message": f"Bad request JSON: {e}"})
    except Exception as e:
        await send({"type": "error", "message": str(e)})


@app.websocket("/ws/manual-scrape")
async def ws_manual_scrape(ws: WebSocket):
    await ws.accept()

    async def send(obj: dict):
        await ws.send_text(json.dumps(obj))

    try:
        raw = await ws.receive_text()
        req = ManualScrapeRequest(**json.loads(raw))

        async def progress(msg: str):
            await send({"type": "progress", "message": msg})

        result = await manual_search_and_scrape_kappal(
            progress_cb=progress,
            headless=False,
            search_timeout=req.search_timeout,
        )

        await send({"type": "complete", "data": result})

    except WebSocketDisconnect:
        pass
    except json.JSONDecodeError as e:
        await send({"type": "error", "message": f"Bad request JSON: {e}"})
    except Exception as e:
        await send({"type": "error", "message": str(e)})


@app.websocket("/ws/batch-scrape")
async def ws_batch_scrape(ws: WebSocket):
    await ws.accept()

    async def send(obj: dict):
        await ws.send_text(json.dumps(obj))

    try:
        raw = await ws.receive_text()
        req = json.loads(raw)
        upload_id = req.get("upload_id")
        workbook_path = BATCH_UPLOADS.get(upload_id)

        # Local-path fallback is useful for developer smoke tests with the
        # generated one-row workbook. The frontend uses upload_id.
        if workbook_path is None and req.get("path"):
            candidate = Path(req["path"])
            if not candidate.is_absolute():
                candidate = Path(__file__).parent / candidate
            workbook_path = candidate

        if workbook_path is None or not workbook_path.exists():
            await send({"type": "error", "message": "Batch workbook not found. Upload it again."})
            return

        async def progress(msg: str):
            await send({"type": "progress", "message": msg})

        result = await run_batch_autosearch(
            workbook_path,
            progress_cb=progress,
            headless=bool(req.get("headless", False)),
            auth_timeout=int(req.get("auth_timeout", 900)),
        )

        await send({"type": "complete", "data": result})

    except WebSocketDisconnect:
        pass
    except json.JSONDecodeError as e:
        await send({"type": "error", "message": f"Bad request JSON: {e}"})
    except Exception as e:
        await send({"type": "error", "message": str(e)})


@app.websocket("/ws/scrape")
async def ws_scrape(ws: WebSocket):
    """
    Protocol (JSON messages both ways):

    Client → Server:
      { "origin": ..., "destination": ..., "cut_off_date": ...,
        "load_type": ..., "quantity": ... }

    Server → Client:
      { "type": "progress", "message": "..." }
      { "type": "complete", "data": { ...results... } }
      { "type": "error",    "message": "..." }
    """
    await ws.accept()

    async def send(obj: dict):
        await ws.send_text(json.dumps(obj))

    try:
        raw = await ws.receive_text()
        req = ScrapeRequest(**json.loads(raw))

        async def progress(msg: str):
            await send({"type": "progress", "message": msg})

        result = await scrape_kappal(
            origin_query     = req.origin,
            destination_query= req.destination,
            cut_off_date     = req.cut_off_date,
            load_type        = req.load_type,
            quantity         = req.quantity,
            origin_service_mode=req.origin_service_mode,
            destination_service_mode=req.destination_service_mode,
            origin_carrier_sd=req.origin_carrier_sd,
            destination_carrier_sd=req.destination_carrier_sd,
            include_nearby_origin=req.include_nearby_origin,
            include_nearby_destination=req.include_nearby_destination,
            charges=req.charges,
            search_reference_name=req.search_reference_name,
            search_currency=req.search_currency,
            progress_cb      = progress,
            headless         = True,
        )

        await send({"type": "complete", "data": result})

    except WebSocketDisconnect:
        pass
    except json.JSONDecodeError as e:
        await send({"type": "error", "message": f"Bad request JSON: {e}"})
    except Exception as e:
        await send({"type": "error", "message": str(e)})


# ─── Run directly ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="0.0.0.0", port=8000, reload=False)
